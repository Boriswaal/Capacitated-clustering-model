import networkx as nx
from networkx.algorithms import approximation
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import utm
from scipy.spatial.distance import euclidean
from sklearn.cluster import KMeans
from itertools import permutations
import math
import numpy as np
from google.cloud import storage
import os


# Load data from Excel files
wb = load_workbook(filename='Hydrogen_transport_data.xlsx')
wb2 = load_workbook(filename='output_data.xlsx')
ws1 = wb["Nodes 2040"]

# Parse node data
nodes = {}
for row in ws1.iter_rows(min_row=2, values_only=True):
    node_name, coordinates, demand, repurposed = row
    lat, lon = map(float, coordinates.split(','))
    nodes[node_name] = {'lat': lat, 'lon': lon, 'demand': demand, 'repurposed': repurposed}

# Create graph
G = nx.complete_graph(list(nodes.keys()))

# Set node attributes
for node, attributes in nodes.items():
    x, y, z, n = utm.from_latlon(attributes['lat'], attributes['lon'])
    G.nodes[node]['x'] = x
    G.nodes[node]['y'] = y

for node, attributes in nodes.items():
    G.nodes[node].update(attributes)

pos = {node: (G.nodes[node]['x'], G.nodes[node]['y']) for node in G.nodes()}

nx.draw_networkx(G, pos)
nx.draw_networkx_nodes(G, pos, nodelist=['P1','P2','P3'], node_color='red')

plt.title("Complete graph")
plt.show()

# Set edge weights
for i, j in G.edges():
    G[i][j]['weight'] = round(euclidean(pos[i], pos[j]))

# Initialize variables
ship_fixed_cost = 0
ship_costkm = 75
ship_capacity = 7285714

# Separate nodes and demands for clustering (excluding ports)
clustering_nodes = [node for node in G.nodes() if node not in {'P1', 'P2', 'P3'}]
x_coord = [G.nodes[node]['x'] for node in clustering_nodes]
y_coord = [G.nodes[node]['y'] for node in clustering_nodes]
demand = [G.nodes[node]['demand'] for node in clustering_nodes]
X = np.array(list(zip(x_coord, y_coord)))

# Function to calculate cluster cost considering multiple ports
def calculate_cluster_cost(G, cluster_nodes, ship_fixed_cost, ship_costkm, ship_capacity):
  ports = ['P1', 'P2', 'P3']
  best_cost = float('inf')
  best_hamilton_cycle = None
  best_port = None

  penalty = 1e18  # Large penalty for exceeding ship capacity

  for port in ports:
      # Create a subgraph including the port and cluster nodes
      subgraph_nodes = [port] + cluster_nodes
      subgraph = G.subgraph(subgraph_nodes)
      
      # Solve the TSP for the subgraph
      cycle = approximation.simulated_annealing_tsp(subgraph, "greedy", source=port)

      # Calculate the cost of the cycle
      try:
          total_distance = sum(G[cycle[i]][cycle[i+1]]['weight'] for i in range(len(cycle) - 1))
      except KeyError as e:
          """print(f"Missing edge in graph: {e}")"""
          continue

      cluster_cost = ship_fixed_cost + (ship_costkm * total_distance / 1000)

      # Apply penalty if the cluster demand exceeds ship capacity
      cluster_demand = sum(demand[node-1] for node in cluster_nodes)
      if cluster_demand > ship_capacity:
          cluster_cost += penalty

      if cluster_cost < best_cost:
          best_cost = cluster_cost
          best_hamilton_cycle = cycle
          best_port = port

  # Return default values if no valid cycle is found
  if best_hamilton_cycle is None:
      return float('inf'), [], None

  return best_cost, best_hamilton_cycle, best_port

# Perform clustering using KMeans++ initialization
num_clusters = math.ceil(sum(demand) / (ship_capacity * 0.8))
clustering = KMeans(n_clusters=num_clusters, init='k-means++', n_init='auto').fit(X)
centers = clustering.cluster_centers_
label = clustering.labels_
u_labels, counts = np.unique(label, return_counts=True)

inertia = sum(euclidean(X[i], centers[label[i]]) ** 2 for i in range(len(X)))
print('Inertia:', inertia)

for i in u_labels:
    plt.scatter(X[label == i, 0], X[label == i, 1], s=10, label=f'Cluster {i}')
plt.legend()
plt.title('Cluster Centers and Nodes')
plt.scatter(centers[:, 0], centers[:, 1], s=30, color='r')
plt.show()

cluster_demand = {}
for i, cluster_label in enumerate(u_labels):
    cluster_demand[cluster_label] = sum(demand[node_idx] for node_idx, cluster_label in enumerate(clustering.labels_) if cluster_label == i)
    cluster_nodes = [node for node, assigned_cluster in zip(clustering_nodes, clustering.labels_) if assigned_cluster == cluster_label]
    print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {cluster_demand[cluster_label]} kg Hydrogen")

# Reassignment based on capacity constraints
cluster_demand = {c: 0 for c in u_labels}
for i, cluster_label in enumerate(label):
    cluster_demand[cluster_label] += demand[i]
    
heap = [(euclidean(X[i], centers[c]), i, c) for c in u_labels for i in range(len(X))]
heap = sorted(heap, reverse=True)
not_connected = list(range(len(X)))
not_full = list(u_labels)
demand_count = {c: 0 for c in list(u_labels)}

centers = np.zeros((num_clusters, X.shape[1]))
for cluster_label in range(num_clusters):
    mask = label == cluster_label
    cluster_indices = np.where(mask)[0]
    if len(cluster_indices) > 0:
        cluster_center = np.mean(X[cluster_indices], axis=0)
        centers[cluster_label] = cluster_center

cluster_assignments = list(clustering.labels_)
while heap:
    d, n, c = heap.pop()
    if c in not_full and n in not_connected:
        if demand_count[c] + demand[n] <= ship_capacity:
            cluster_assignments[n] = c
            not_connected.remove(n)
            demand_count[c] += demand[n]
            if demand_count[c] >= ship_capacity:
                not_full.remove(c)
        else:
            closest_cluster = min(not_full, key=lambda x: euclidean(X[n], centers[x]))
            if demand_count[closest_cluster] + demand[n] <= ship_capacity:
                cluster_assignments[n] = closest_cluster
                not_connected.remove(n)
                demand_count[closest_cluster] += demand[n]
                if demand_count[closest_cluster] >= ship_capacity:
                    not_full.remove(closest_cluster)

for i in u_labels:
    plt.scatter(X[label == i, 0], X[label == i, 1], s=10, label=f'Cluster {i}')
plt.legend()
plt.title('Cluster Centers and Nodes')
plt.scatter(centers[:, 0], centers[:, 1], s=30, color='r')
plt.show()

label = np.array(cluster_assignments)
for cluster_label in range(num_clusters):
    cluster_nodes = [node for node, assigned_cluster in zip(clustering_nodes, label) if assigned_cluster == cluster_label]
    cluster_demand[cluster_label] = sum(demand[node_idx] for node_idx, assigned_cluster in enumerate(label) if assigned_cluster == cluster_label)
    print(f"Improved cluster {cluster_label} nodes: {cluster_nodes} \nImproved cluster {cluster_label} demand: {cluster_demand[cluster_label]} kg Hydrogen")

new_inertia = sum(euclidean(X[i], centers[label[i]]) ** 2 for i in range(len(X)))
print('New Inertia:', new_inertia)

for cluster_label in range(num_clusters):
    mask = label == cluster_label
    cluster_indices = np.where(mask)[0]
    if len(cluster_indices) > 0:
        cluster_center = np.mean(X[cluster_indices], axis=0)
        centers[cluster_label] = cluster_center

def optimize_clusters(label, centers, X, cluster_demand, ship_capacity):
    swaps = 0
    improved = True
    node_list = clustering_nodes  # Update to use clustering nodes
    found_valid_assignment = False
    num_clusters = len(centers)

    while improved:
        improved = False

        # Step 1: Check if moving a single node to another cluster reduces the total cost
        for i in range(len(X)):
            for j in range(num_clusters):
                if label[i] != j:
                    cluster_i = label[i]
                    cluster_j = j

                    new_demand_i = cluster_demand[cluster_i] - demand[i]
                    new_demand_j = cluster_demand[cluster_j] + demand[i]

                    if new_demand_i <= ship_capacity and new_demand_j <= ship_capacity:
                        temp_label = label.copy()
                        temp_label[i] = j

                        cluster_i_nodes_before = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_i]
                        cluster_j_nodes_before = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_j]
                        cluster_i_nodes_after = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_i]
                        cluster_j_nodes_after = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_j]

                        original_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes_before, ship_fixed_cost, ship_costkm, ship_capacity)
                        original_cluster_j_cost, _, _= calculate_cluster_cost(G, cluster_j_nodes_before, ship_fixed_cost, ship_costkm, ship_capacity)
                        new_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes_after, ship_fixed_cost, ship_costkm, ship_capacity)
                        new_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes_after, ship_fixed_cost, ship_costkm, ship_capacity)

                        if (new_cluster_i_cost + new_cluster_j_cost) < (original_cluster_i_cost + original_cluster_j_cost):
                            label[i] = j
                            cluster_demand[cluster_i] = new_demand_i
                            cluster_demand[cluster_j] = new_demand_j
                            swaps += 1
                            improved = True

                            print(f"Moving node {node_list[i]} from cluster {cluster_i} to cluster {cluster_j}")
                            print(f"Current clusters after move {swaps}:")
                            for cluster_label in range(num_clusters):
                                cluster_nodes = [node for node, assigned_cluster in zip(node_list, label) if assigned_cluster == cluster_label]
                                print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {sum(G.nodes[node]['demand'] for node in cluster_nodes)} kg Hydrogen")

                            # Re-calculate cluster centers
                            for cluster_label in range(num_clusters):
                                mask = label == cluster_label
                                cluster_indices = np.where(mask)[0]
                                if len(cluster_indices) > 0:
                                    cluster_center = np.mean(X[cluster_indices], axis=0)
                                    centers[cluster_label] = cluster_center

        # Step 2: Check if swapping single nodes from clusters reduces the total cost
        for i, j in permutations(range(len(X)), 2):
            if label[i] != label[j]:
                cluster_i = label[i]
                cluster_j = label[j]

                new_demand_i = cluster_demand[cluster_i] - demand[i] + demand[j]
                new_demand_j = cluster_demand[cluster_j] - demand[j] + demand[i]

                if new_demand_i <= ship_capacity and new_demand_j <= ship_capacity:
                    temp_label = label.copy()
                    temp_label[i], temp_label[j] = temp_label[j], temp_label[i]

                    cluster_i_nodes_before = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_i]
                    cluster_j_nodes_before = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_j]
                    cluster_i_nodes_after = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_i]
                    cluster_j_nodes_after = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_j]

                    original_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes_before, ship_fixed_cost, ship_costkm, ship_capacity)
                    original_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes_before, ship_fixed_cost, ship_costkm, ship_capacity)
                    new_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes_after, ship_fixed_cost, ship_costkm, ship_capacity)
                    new_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes_after, ship_fixed_cost, ship_costkm, ship_capacity)

                    if (new_cluster_i_cost + new_cluster_j_cost) < (original_cluster_i_cost + original_cluster_j_cost):
                        label[i], label[j] = label[j], label[i]
                        cluster_demand[cluster_i] = new_demand_i
                        cluster_demand[cluster_j] = new_demand_j
                        swaps += 1
                        improved = True

                        print(f"Swapping node {node_list[i]} from cluster {cluster_i} with node {node_list[j]} from cluster {cluster_j}")
                        print(f"Current clusters after swap {swaps}:")
                        for cluster_label in range(num_clusters):
                            cluster_nodes = [node for node, assigned_cluster in zip(node_list, label) if assigned_cluster == cluster_label]
                            print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {sum(G.nodes[node]['demand'] for node in cluster_nodes)} kg Hydrogen")

                        # Re-calculate cluster centers
                        for cluster_label in range(num_clusters):
                            mask = label == cluster_label
                            cluster_indices = np.where(mask)[0]
                            if len(cluster_indices) > 0:
                                cluster_center = np.mean(X[cluster_indices], axis=0)
                                centers[cluster_label] = cluster_center

        # Step 3: Check if swapping multiple nodes from cluster_i with multiple nodes from cluster_j reduces the cost
        for cluster_i in range(num_clusters):
            for cluster_j in range(num_clusters):
                if cluster_i != cluster_j:
                    cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_i]
                    cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_j]

                    for comb_i in permutations(cluster_i_nodes, 2):  # Adjust the range as needed
                        for comb_j in permutations(cluster_j_nodes, 2):  # Adjust the range as needed
                            new_demand_i = cluster_demand[cluster_i] - sum(G.nodes[node]['demand'] for node in comb_i) + sum(G.nodes[node]['demand'] for node in comb_j)
                            new_demand_j = cluster_demand[cluster_j] - sum(G.nodes[node]['demand'] for node in comb_j) + sum(G.nodes[node]['demand'] for node in comb_i)
                            if new_demand_i <= ship_capacity and new_demand_j <= ship_capacity:
                                temp_label = label.copy()
                                for node in comb_i:
                                    temp_label[node_list.index(node)] = cluster_j
                                for node in comb_j:
                                    temp_label[node_list.index(node)] = cluster_i
                                new_cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_i]
                                new_cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_j]
                                new_cluster_i_cost, _, _ = calculate_cluster_cost(G, new_cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                new_cluster_j_cost, _, _ = calculate_cluster_cost(G, new_cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                original_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                original_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                if (new_cluster_i_cost + new_cluster_j_cost) < (original_cluster_i_cost + original_cluster_j_cost):
                                    label = temp_label
                                    cluster_demand[cluster_i] = new_demand_i
                                    cluster_demand[cluster_j] = new_demand_j
                                    swaps += 1
                                    improved = True
                                    print(f"Swapping nodes {comb_i} from cluster {cluster_i} with nodes {comb_j} from cluster {cluster_j}")
                                    print(f"Current clusters after swap {swaps}:")
                                    for cluster_label in range(num_clusters):
                                        cluster_nodes = [node for node, assigned_cluster in zip(node_list, label) if assigned_cluster == cluster_label]
                                        print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {sum(G.nodes[node]['demand'] for node in cluster_nodes)} kg Hydrogen")

                                    # Re-calculate cluster centers
                                    for cluster_label in range(num_clusters):
                                        mask = label == cluster_label
                                        cluster_indices = np.where(mask)[0]
                                        if len(cluster_indices) > 0:
                                            cluster_center = np.mean(X[cluster_indices], axis=0)
                                            centers[cluster_label] = cluster_center
                                if improved:
                                    break
                        if improved:
                            break
                    if improved:
                        break

        # Step 4: Check if swapping a single node from cluster_i with multiple nodes from cluster_j reduces the cost
        for cluster_i in range(num_clusters):
            for cluster_j in range(num_clusters):
                if cluster_i != cluster_j:
                    cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_i]
                    cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_j]

                    for node_i in cluster_i_nodes:
                        for k in range(1, len(cluster_j_nodes) + 1):
                            for swap_comb in permutations(cluster_j_nodes, k):
                                new_demand_i = cluster_demand[cluster_i] - G.nodes[node_i]['demand'] + sum(G.nodes[node]['demand'] for node in swap_comb)
                                new_demand_j = cluster_demand[cluster_j] - sum(G.nodes[node]['demand'] for node in swap_comb) + G.nodes[node_i]['demand']
                                if new_demand_i <= ship_capacity and new_demand_j <= ship_capacity:
                                    temp_label = label.copy()
                                    temp_label[node_list.index(node_i)] = cluster_j
                                    for node in swap_comb:
                                        temp_label[node_list.index(node)] = cluster_i
                                    new_cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_i]
                                    new_cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_j]
                                    new_cluster_i_cost, _, _ = calculate_cluster_cost(G, new_cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                    new_cluster_j_cost, _, _ = calculate_cluster_cost(G, new_cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                    original_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                    original_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                    if (new_cluster_i_cost + new_cluster_j_cost) < (original_cluster_i_cost + original_cluster_j_cost):
                                        label = temp_label
                                        cluster_demand[cluster_i] = new_demand_i
                                        cluster_demand[cluster_j] = new_demand_j
                                        swaps += 1
                                        improved = True
                                        print(f"Swapping node {node_i} from cluster {cluster_i} with nodes {swap_comb} from cluster {cluster_j}")
                                        print(f"Current clusters after swap {swaps}:")
                                        for cluster_label in range(num_clusters):
                                            cluster_nodes = [node for node, assigned_cluster in zip(node_list, label) if assigned_cluster == cluster_label]
                                            print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {sum(G.nodes[node]['demand'] for node in cluster_nodes)} kg Hydrogen")

                                        # Re-calculate cluster centers
                                        for cluster_label in range(num_clusters):
                                            mask = label == cluster_label
                                            cluster_indices = np.where(mask)[0]
                                            if len(cluster_indices) > 0:
                                                cluster_center = np.mean(X[cluster_indices], axis=0)
                                                centers[cluster_label] = cluster_center
                                if improved:
                                    break
                            if improved:
                                break
                        if improved:
                            break
                    if improved:
                        break

        # Step 5: Check if swapping multiple nodes from cluster_i with a single node from cluster_j reduces the cost
        for cluster_i in range(num_clusters):
            for cluster_j in range(num_clusters):
                if cluster_i != cluster_j:
                    cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_i]
                    cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if label[node_idx] == cluster_j]

                    for swap_comb in permutations(cluster_i_nodes, 2):  # Adjust the range as needed
                        for node_j in cluster_j_nodes:
                            new_demand_i = cluster_demand[cluster_i] - sum(G.nodes[node]['demand'] for node in swap_comb) + G.nodes[node_j]['demand']
                            new_demand_j = cluster_demand[cluster_j] - G.nodes[node_j]['demand'] + sum(G.nodes[node]['demand'] for node in swap_comb)
                            if new_demand_i <= ship_capacity and new_demand_j <= ship_capacity:
                                temp_label = label.copy()
                                temp_label[node_list.index(node_j)] = cluster_i
                                for node in swap_comb:
                                    temp_label[node_list.index(node)] = cluster_j
                                new_cluster_i_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_i]
                                new_cluster_j_nodes = [node for node_idx, node in enumerate(node_list) if temp_label[node_idx] == cluster_j]
                                new_cluster_i_cost, _, _ = calculate_cluster_cost(G, new_cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                new_cluster_j_cost, _, _ = calculate_cluster_cost(G, new_cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                original_cluster_i_cost, _, _ = calculate_cluster_cost(G, cluster_i_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                original_cluster_j_cost, _, _ = calculate_cluster_cost(G, cluster_j_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
                                if (new_cluster_i_cost + new_cluster_j_cost) < (original_cluster_i_cost + original_cluster_j_cost):
                                    label = temp_label
                                    cluster_demand[cluster_i] = new_demand_i
                                    cluster_demand[cluster_j] = new_demand_j
                                    swaps += 1
                                    improved = True
                                    print(f"Swapping nodes {swap_comb} from cluster {cluster_i} with node {node_j} from cluster {cluster_j}")
                                    print(f"Current clusters after swap {swaps}:")
                                    for cluster_label in range(num_clusters):
                                        cluster_nodes = [node for node, assigned_cluster in zip(node_list, label) if assigned_cluster == cluster_label]
                                        print(f"Cluster {cluster_label} nodes: {cluster_nodes} \nCluster {cluster_label} demand: {sum(G.nodes[node]['demand'] for node in cluster_nodes)} kg Hydrogen")

                                    # Re-calculate cluster centers
                                    for cluster_label in range(num_clusters):
                                        mask = label == cluster_label
                                        cluster_indices = np.where(mask)[0]
                                        if len(cluster_indices) > 0:
                                            cluster_center = np.mean(X[cluster_indices], axis=0)
                                            centers[cluster_label] = cluster_center
                                if improved:
                                    break
                        if improved:
                            break
                    if improved:
                        break

        print("Demand count before handling over-capacity clusters:", cluster_demand)
        over_capacity_clusters = [cluster for cluster, total_demand in cluster_demand.items() if total_demand > ship_capacity]
        print("Over capacity clusters:", over_capacity_clusters)
        if not over_capacity_clusters:
            found_valid_assignment = True
            break  # Exit the loop if no over-capacity clusters are found

    # Increase the number of clusters by 1 if there are over-capacity clusters
    if over_capacity_clusters:
        print("Increasing the number of clusters by 1 due to over-capacity clusters.")
        num_clusters += 1
        new_cluster_center = np.mean(X, axis=0)  # Initialize the new cluster center to the mean of all points
        centers = np.vstack([centers, new_cluster_center])  # Add the new cluster center to the list of centers

        # Re-run KMeans with the new number of clusters
        kmeans = KMeans(n_clusters=num_clusters, init=centers, n_init=1)
        new_labels = kmeans.fit_predict(X)

        # Recalculate cluster demands
        new_cluster_demand = {i: 0 for i in range(num_clusters)}
        for i, demand_value in enumerate(demand):
            new_cluster_demand[new_labels[i]] += demand_value

        # Update the labels and cluster demands
        label = new_labels
        cluster_demand = new_cluster_demand

        # Re-optimize the clusters with the new assignments
        return optimize_clusters(label, centers, X, cluster_demand, ship_capacity)

    return swaps, found_valid_assignment, label, cluster_demand


# Check if all nodes are in the closest cluster
all_in_closest_cluster = True
for i, assignment in enumerate(cluster_assignments):
    closest_cluster = min(range(len(centers)), key=lambda x: euclidean(X[i], centers[x]))
    if assignment != closest_cluster:
        all_in_closest_cluster = False
        break

# Check for over-capacity clusters
over_capacity_clusters = [cluster for cluster, total_demand in cluster_demand.items() if total_demand > ship_capacity]

# If all nodes are in their closest cluster and there are no over-capacity clusters, calculate the cost directly
if all_in_closest_cluster and not over_capacity_clusters:
    cluster_costs = []
    for c in range(len(centers)):
        cluster_nodes = [i for i in range(len(cluster_assignments)) if cluster_assignments[i] == c]
        cost, _, _ = calculate_cluster_cost(G, cluster_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
        cluster_costs.append(cost)
    total_cost = sum(cluster_costs)
    print(f"All nodes are in their closest cluster. Total cost: {total_cost}")
else:
    # Perform node swaps to further optimize the cluster assignments
    swaps, found_valid_assignment, optimized_label, optimized_cluster_demand = optimize_clusters(
        label, centers, X, cluster_demand, ship_capacity
    )
    print('Number of swaps:', swaps)


final_label = label.copy()
final_cluster_demand = cluster_demand.copy()

# Update cluster demands based on final labels
final_cluster_demand = {label_idx: sum(demand[node_idx] for node_idx in range(len(final_label)) if final_label[node_idx] == label_idx) for label_idx in np.unique(final_label)}

#Recreate cluster centers based on final assignments
num_clusters = len(np.unique(final_label))
final_centers = np.zeros((num_clusters, X.shape[1]))
for cluster_label in range(num_clusters):
    mask = final_label == cluster_label
    cluster_indices = np.where(mask)[0]
    if len(cluster_indices) > 0:
        cluster_center = np.mean(X[cluster_indices], axis=0)
        final_centers[cluster_label] = cluster_center

for cluster_label in range(num_clusters):
    cluster_nodes = [node for node, assigned_cluster in zip(clustering_nodes, final_label) if assigned_cluster == cluster_label]
    print(f"Final cluster {cluster_label} nodes: {cluster_nodes}")

# Plot each cluster's nodes and center
for cluster_label in range(num_clusters):
    cluster_nodes = X[final_label == cluster_label]
    cluster_center = final_centers[cluster_label]
    plt.scatter(cluster_nodes[:, 0], cluster_nodes[:, 1], s=10, label=f'Cluster {cluster_label} nodes')
    plt.legend()
    plt.scatter(cluster_center[0], cluster_center[1], s=30, color='red')
plt.title('Final Cluster Centers and Nodes')
plt.show()

# Final cluster assignments and costs
cluster_costs = {}
best_hamilton_paths = {}
best_ports = {}

for cluster_label in range(num_clusters):
    cluster_nodes = [node for node, assigned_cluster in zip(clustering_nodes, final_label) if assigned_cluster == cluster_label]
    cluster_cost, hamilton_path, best_port = calculate_cluster_cost(G, cluster_nodes, ship_fixed_cost, ship_costkm, ship_capacity)
    cluster_costs[cluster_label] = cluster_cost
    best_hamilton_paths[cluster_label] = hamilton_path
    best_ports[cluster_label] = best_port
   
    cluster_demand = sum(demand[node_idx] for node_idx, assigned_cluster in enumerate(final_label) if assigned_cluster == cluster_label)
    print(f"Final cluster {cluster_label} nodes: {cluster_nodes} \nFinal cluster {cluster_label} demand: {cluster_demand} kg Hydrogen\nCost: €{cluster_cost} using port {best_port}")

new_inertia = sum(euclidean(X[i], final_centers[final_label[i]]) ** 2 for i in range(len(X)))
print('Final Inertia:', new_inertia)

# Plot the best Hamilton paths for the clusters
plt.figure(figsize=(12, 8))
for label, path in best_hamilton_paths.items():
    subgraph = G.subgraph(path)
    nx.draw_networkx(subgraph, pos, node_color='lightblue', with_labels=True)
    nx.draw_networkx_nodes(subgraph, pos, nodelist=[best_ports[label]], node_color='red')
    nx.draw_networkx_edges(subgraph, pos, edgelist=[(path[i], path[i+1]) for i in range(len(path) - 1)], edge_color='blue', width=2)
    labels = nx.get_edge_attributes(subgraph, 'weight')
    nx.draw_networkx_edge_labels(subgraph, pos, edge_labels=labels, font_size=8, label_pos=0.5)
    total_distance = sum(subgraph[path[i]][path[i+1]]['weight'] for i in range(len(path) - 1))
    plt.title

# Plot each cluster separately with the best Hamilton path
for label, path in best_hamilton_paths.items():
    plt.figure(figsize=(12, 8))
    subgraph = G.subgraph(path)
    nx.draw_networkx(subgraph, pos, node_color='lightblue', with_labels=True, node_size=500, font_size=10)
    nx.draw_networkx_nodes(subgraph, pos, nodelist=[best_ports[label]], node_color='red', node_size=700)
    nx.draw_networkx_edges(subgraph, pos, edgelist=[(path[i], path[i+1]) for i in range(len(path) - 1)], edge_color='blue', width=2)
    labels = nx.get_edge_attributes(subgraph, 'weight')
    nx.draw_networkx_edge_labels(subgraph, pos, edge_labels=labels, font_size=8, label_pos=0.5)
    total_distance = sum(subgraph[path[i]][path[i+1]]['weight'] for i in range(len(path) - 1))
    plt.title(f'Best Hamilton Path for Cluster {label} using port {best_ports[label]} - Distance: {total_distance}')
    plt.show()

def print_hamiltonian_path_edges(G, hamilton_path):
    for i in range(len(hamilton_path) - 1):
        node1 = hamilton_path[i]
        node2 = hamilton_path[i + 1]
        distance = G[node1][node2]['weight']
        print(f"Edge: {node1} -> {node2}, Distance: {distance} meters")

# Print the edges and distances of the best Hamiltonian paths for each cluster
for cluster_label, hamilton_path in best_hamilton_paths.items():
    print(f"\nBest Hamiltonian Path for Cluster {cluster_label}:")
    print_hamiltonian_path_edges(G, hamilton_path)
    total_distance = sum(G[hamilton_path[i]][hamilton_path[i+1]]['weight'] for i in range(len(hamilton_path) - 1))
    print(f"Total Distance for Cluster {cluster_label}: {total_distance} meters")

def save_cluster_results():
    # Create a new sheet to store the results
    output_wb = wb2
    output_ws = output_wb.create_sheet(title='Cluster Results')

    # Write the headers
    headers = ['Cluster', 'Nodes', 'Total Demand (kg Hydrogen)', 'Hamiltonian Path Edges', 'Total Distance (meters)', 'Total Cost (€)']
    output_ws.append(headers)

    # Function to save plots
    def save_plot(cluster_label, path, best_port, total_distance):
        plt.figure(figsize=(12, 8))
        subgraph = G.subgraph(path)
        nx.draw_networkx(subgraph, pos, node_color='lightblue', with_labels=True, node_size=500, font_size=10)
        nx.draw_networkx_nodes(subgraph, pos, nodelist=[best_port], node_color='red', node_size=700)
        nx.draw_networkx_edges(subgraph, pos, edgelist=[(path[i], path[i+1]) for i in range(len(path) - 1)], edge_color='blue', width=2)
        labels = nx.get_edge_attributes(subgraph, 'weight')
        nx.draw_networkx_edge_labels(subgraph, pos, edge_labels=labels, font_size=8, label_pos=0.5)
        plt.title(f'Best Hamilton Path for Cluster {cluster_label} using start port {best_port} - Distance: {total_distance}')
        file_path = f'cluster_{cluster_label}_path.png'
        plt.savefig(file_path)
        plt.close()
        return file_path

    for cluster_label, hamilton_path in best_hamilton_paths.items():
        cluster_nodes = [node for node, assigned_cluster in zip(clustering_nodes, final_label) if assigned_cluster == cluster_label]
        cluster_demand = sum(demand[node_idx] for node_idx, assigned_cluster in enumerate(final_label) if assigned_cluster == cluster_label)
        total_distance = sum(G[hamilton_path[i]][hamilton_path[i+1]]['weight'] for i in range(len(hamilton_path) - 1))
        total_cost = cluster_costs[cluster_label]
        
        # Format the Hamiltonian path edges as a string
        edges = [f"{hamilton_path[i]} -> {hamilton_path[i+1]} ({G[hamilton_path[i]][hamilton_path[i+1]]['weight']} meters)" for i in range(len(hamilton_path) - 1)]
        edges_str = '; '.join(edges)
        
        # Save the plot for the current cluster
        plot_file = save_plot(cluster_label, hamilton_path, best_ports[cluster_label], total_distance)

        # Write the data to the sheet
        output_ws.append([cluster_label, ', '.join(map(str, cluster_nodes)), cluster_demand, edges_str, total_distance, total_cost])

        # Insert the plot image into the sheet
        img = Image(plot_file)
        output_ws.add_image(img, f'H{cluster_label + 2}')  # Adjust the cell location as needed

    # Save the workbook locally
    output_excel_path = 'output_data.xlsx'
    output_wb.save(output_excel_path)


# Call the function to execute the workflow
save_cluster_results()